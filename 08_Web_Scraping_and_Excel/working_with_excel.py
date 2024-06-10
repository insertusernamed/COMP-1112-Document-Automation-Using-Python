import openpyxl
import os

# Changing the working directory to the current file's directory so that the excel files will be created in the same directory as the script
os.chdir(os.path.dirname(os.path.abspath(__file__)))

# Creating a new Excel workbook and sheet
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Data"

# Writing data to the Excel sheet
data = [
    ["Name", "Age", "City"],
    ["Daniel", 19, "Cookstown"],
    ["Jenna", 19, "Barrie"],
    ["Christian", 24, "Beeton"],
]

for row in data:
    sheet.append(row)

# Saving the workbook
workbook.save("example.xlsx")

# Reading data from an Excel file
workbook = openpyxl.load_workbook("example.xlsx")
sheet = workbook["Data"]

# Printing the contents of the sheet
for row in sheet.iter_rows(values_only=True):
    print(row)

# Modifying data in the Excel sheet
sheet["B2"] = 20  # Change Daniel's age from 19 to 20

# Saving the modified workbook
workbook.save("example_modified.xlsx")
